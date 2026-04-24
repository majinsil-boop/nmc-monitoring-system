"""

국회의안정보시스템 자동 검색 스크립트

키워드: 응급의료, 119구조구급법, 공공보건의료법, 의료법, 대안(보건복지위원장), 필수의료법, 의료, 응급

"""



import asyncio

import json

import csv

import os

import re

from datetime import datetime, timedelta

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout



import subprocess



try:

    import openpyxl

    from openpyxl.styles import PatternFill, Font, Alignment

    _XLSX_AVAILABLE = True

except ImportError:

    _XLSX_AVAILABLE = False

    print("[경고] openpyxl 미설치 - history_db.xlsx 기능 비활성화. pip install openpyxl")





def _summarize_with_claude(raw_text: str, bill_name: str = "") -> str:

    """

    Claude Code CLI(claude -p)로 의안 내용을 2-3문장 한국어 요약.

    실패 시 원문 그대로 반환.

    """

    if not raw_text or len(raw_text) < 30:

        return raw_text



    prompt = (

        "다음은 국회 의안의 제안이유 및 주요내용입니다.\n"

        "핵심 개정 사유와 주요 변경 내용을 2~3문장으로 간결하게 한국어로 요약해 주세요.\n"

        "전문 용어는 그대로 유지하되 누구나 이해하기 쉽게 써 주세요.\n"

        "요약문만 출력하고 다른 설명은 붙이지 마세요.\n\n"

        f"의안명: {bill_name}\n\n"

        f"내용:\n{raw_text[:3000]}"

    )



    import tempfile

    tmp_path = None

    try:

        with tempfile.NamedTemporaryFile(

            mode="w", suffix=".txt", encoding="utf-8", delete=False

        ) as tmp:

            tmp.write(prompt)

            tmp_path = tmp.name



        ps_cmd = (

            "[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; "

            "$OutputEncoding = [System.Text.Encoding]::UTF8; "

            f"$p = Get-Content -Raw -Encoding UTF8 '{tmp_path}'; "

            "claude -p $p"

        )

        result = subprocess.run(

            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_cmd],

            capture_output=True,

            timeout=90,

        )

        try:

            stdout = result.stdout.decode("utf-8").strip()

        except UnicodeDecodeError:

            stdout = result.stdout.decode("cp949", errors="replace").strip()

        output = stdout



        if result.returncode == 0 and output:

            return output

        if result.stderr:

            print(f"  [Claude 요약 오류] {result.stderr[:120]}")

        return raw_text

    except subprocess.TimeoutExpired:

        print("  [Claude 요약 타임아웃]")

        return raw_text

    except Exception as e:

        print(f"  [Claude 요약 실패] {e}")

        return raw_text

    finally:

        if tmp_path and os.path.exists(tmp_path):

            os.unlink(tmp_path)



BASE_URL = "https://likms.assembly.go.kr/bill/bi/main/mainPage.do"

DETAIL_BASE = "https://likms.assembly.go.kr/bill/bi/billDetailPage.do"



KEYWORDS = [

    "응급의료",

    "119구조구급",

    "공공보건의료법",

    "의료법",

    "필수의료",

    "대안",

    "의료",

    "응급",

]



# 의료법·119구조구급·필수의료·대안·의료: 응급 관련 내용 포함 필수

_NARROW_KEYWORDS: set[str] = {"의료법", "119구조구급", "필수의료", "대안", "의료"}

_NARROW_CONTENT_FILTER: list[str] = [

    "응급실", "응급의료", "중증외상", "구급", "응급처치",

]

# 응급 키워드: 법안명 또는 요약에 "응급" 포함 필수

_EMERGENCY_ONLY_KEYWORDS: set[str] = {"응급"}



# 우선순위 의원 목록 — 발의자에 포함 시 ★ 표시 및 최상단 배치

PRIORITY_MEMBERS: set[str] = {"이수진", "김미애", "김윤", "김선민", "한지아", "이주영"}



HISTORY_DB = "history_db.xlsx"



RECENT_DAYS  = 7

STATUS_DAYS  = 7

SEARCH_CONCURRENCY = 4   # 동시 키워드 검색 수

DETAIL_CONCURRENCY = 5   # 동시 상세 페이지 수집 수



_DONE_KEYWORDS = ("폐기", "공포", "이송", "가결", "부결", "철회")





# ── 이력 DB (history_db.xlsx) ────────────────────────────────────────────────



def _load_history() -> dict[str, dict]:

    """history_db.xlsx에서 이전 수집 이력을 로드. {bill_no: row_dict}"""

    if not _XLSX_AVAILABLE or not os.path.exists(HISTORY_DB):

        return {}

    try:

        wb = openpyxl.load_workbook(HISTORY_DB)

        ws = wb.active

        headers = [c.value for c in ws[1]]

        history: dict[str, dict] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):

            d = dict(zip(headers, row))

            bill_no = str(d.get("bill_no", "")).strip()

            if bill_no:

                history[bill_no] = d

        return history

    except Exception as e:

        print(f"[이력 DB 로드 오류] {e}")

        return {}





def _save_history(bills: list[dict], history: dict[str, dict]):

    """현재 수집 결과를 history_db.xlsx에 누적 저장한다."""

    if not _XLSX_AVAILABLE:

        return

    today_str = datetime.now().strftime("%Y-%m-%d %H:%M")



    # 기존 이력을 현재 결과로 업데이트

    for bill in bills:

        bill_no = str(bill.get("bill_no", "")).strip()

        if not bill_no:

            continue

        prev = history.get(bill_no, {})

        history[bill_no] = {

            "bill_no":              bill_no,

            "bill_name":            bill.get("bill_name", ""),

            "proposer":             bill.get("proposer", ""),

            "proposed_date":        bill.get("proposed_date", ""),

            "vote_date":            bill.get("vote_date", ""),

            "status":               bill.get("status", ""),

            "status_changed":       bill.get("status_changed", False),

            "status_changed_date":  bill.get("status_changed_date", ""),

            "prev_status":          prev.get("status", ""),

            "is_key_member":        bill.get("is_key_member", False),

            "legislative_notice":   bill.get("legislative_notice", ""),

            "summary":              bill.get("summary", ""),

            "keyword":              bill.get("keyword", ""),

            "url":                  bill.get("url", ""),

            "first_seen":           prev.get("first_seen") or today_str,

            "last_updated":         today_str,

        }



    if not history:

        return



    COLS = [

        "bill_no", "bill_name", "proposer", "proposed_date", "vote_date",

        "status", "status_changed", "status_changed_date", "prev_status",

        "is_key_member", "legislative_notice", "summary", "keyword", "url",

        "first_seen", "last_updated",

    ]



    try:

        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "의안이력"



        # 헤더

        header_fill = PatternFill("solid", fgColor="4472C4")

        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(COLS, 1):

            cell = ws.cell(row=1, column=col_idx, value=col_name)

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(horizontal="center")



        # 데이터 행

        changed_fill  = PatternFill("solid", fgColor="FFE699")  # 상태변경 행: 노란색

        priority_fill = PatternFill("solid", fgColor="C6EFCE")  # 우선의원 행: 연두색



        for row_idx, row_dict in enumerate(history.values(), 2):

            prev_s    = str(row_dict.get("prev_status") or "").strip()

            curr_s    = str(row_dict.get("status") or "").strip()

            is_changed  = bool(prev_s and prev_s != curr_s)

            proposer  = str(row_dict.get("proposer") or "")

            is_priority = any(m in proposer for m in PRIORITY_MEMBERS)



            for col_idx, col_name in enumerate(COLS, 1):

                val = row_dict.get(col_name, "")

                cell = ws.cell(row=row_idx, column=col_idx, value=val)

                if is_changed:

                    cell.fill = changed_fill

                elif is_priority:

                    cell.fill = priority_fill



        # 열 너비 자동 조정

        col_widths = {

            "bill_no": 14, "bill_name": 50, "proposer": 20,

            "proposed_date": 14, "vote_date": 14, "status": 20,

            "status_changed": 14, "status_changed_date": 18, "prev_status": 20,

            "is_key_member": 14, "legislative_notice": 30, "summary": 60,

            "keyword": 20, "url": 40, "first_seen": 18, "last_updated": 18,

        }

        for col_idx, col_name in enumerate(COLS, 1):

            ws.column_dimensions[

                openpyxl.utils.get_column_letter(col_idx)

            ].width = col_widths.get(col_name, 15)



        wb.save(HISTORY_DB)

        print(f"[이력 DB] {HISTORY_DB} 저장 완료 (총 {len(history)}건)")

    except Exception as e:

        print(f"[이력 DB 저장 오류] {e}")





# ── 필터 헬퍼 ────────────────────────────────────────────────────────────────



_CORE_FILTER_KW: list[str] = ["응급실", "응급의료", "중증외상", "중증응급"]


def _passes_content_filter(bill: dict) -> bool:
    """
    합격 기준:
      1) 의안명이 '응급의료'로 시작 → 합격
      2) 의안명+요약에 응급실·응급의료·중증외상·중증응급 중 하나 포함 → 합격
      3) 소관위원회가 보건복지위원회 AND 입법예고 기간 중 → 합격
    """
    name = bill.get("bill_name", "")
    text = name + " " + bill.get("summary", "")
    committee = bill.get("committee", "")

    if "응급의료" in name:
        return True

    if any(kw in text for kw in _CORE_FILTER_KW):
        return True

    # 보건복지위원회 소관 + 입법예고 기간 중인 의안
    if "보건복지" in committee and _is_notice_active(bill.get("legislative_notice", "")):
        print(f"    → [보건복지위] 입법예고 중 의안 포함: {name[:40]}")
        return True

    return False





def _is_active(status: str) -> bool:

    return not any(kw in status for kw in _DONE_KEYWORDS)





def _is_recent(bill: dict, days: int = RECENT_DAYS) -> bool:

    cutoff = datetime.now() - timedelta(days=days)

    for field in ("proposed_date", "vote_date"):

        raw = bill.get(field, "").strip()

        if not raw:

            continue

        try:

            normalized = raw.replace(".", "-").replace("/", "-")

            if datetime.strptime(normalized[:10], "%Y-%m-%d") >= cutoff:

                return True

        except ValueError:

            pass

    return False





def _is_recent_status(status_date: str, days: int = STATUS_DAYS) -> bool:

    if not status_date:

        return False

    cutoff = datetime.now() - timedelta(days=days)

    try:

        normalized = status_date.replace(".", "-").replace("/", "-")

        return datetime.strptime(normalized[:10], "%Y-%m-%d") >= cutoff

    except ValueError:

        return False





def _is_notice_active(notice: str) -> bool:

    if not notice:

        return False

    m = re.search(r"~\s*(\d{4}[.\-]\d{2}[.\-]\d{2})", notice)

    if not m:

        return False

    try:

        normalized = m.group(1).replace(".", "-")

        end_date = datetime.strptime(normalized, "%Y-%m-%d")

        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        return end_date >= today

    except ValueError:

        return False





def _is_priority(bill: dict) -> bool:

    """발의자 중 우선순위 의원이 포함되어 있으면 True."""

    proposer = bill.get("proposer", "")

    return any(m in proposer for m in PRIORITY_MEMBERS)





def _clean_bill_name(title_attr: str) -> str:

    return re.sub(r"\s*\([^)]*의안[^)]*\)\s*$", "", title_attr).strip()





def _apply_status_change_label(bill: dict, history: dict[str, dict]):

    """이전 상태와 비교해 상태변경 레이블을 bill에 추가한다."""

    bill_no = str(bill.get("bill_no", "")).strip()

    prev = history.get(bill_no, {})

    prev_status = str(prev.get("status", "")).strip()

    curr_status = str(bill.get("status", "")).strip()



    if prev_status and prev_status != curr_status:

        bill["prev_status"] = prev_status

        bill["status_change_label"] = f"[상태변경: {prev_status} → {curr_status}]"

        bill["status_changed"] = True

        print(f"    ★ 상태변경 감지: {bill.get('bill_name','')[:30]} "

              f"| {prev_status} → {curr_status}")

    else:

        bill["prev_status"] = prev_status

        bill["status_change_label"] = ""

        bill["status_changed"] = False





# ── 상세 페이지 수집 ─────────────────────────────────────────────────────────



async def _fetch_bill_detail(page, url: str) -> tuple[str, str, str, str]:
    """(summary, status_date, notice, committee) 반환."""

    if not url:

        return "", "", "", ""

    try:

        await page.goto(url, wait_until="domcontentloaded", timeout=25000)

        await asyncio.sleep(0.3)

        await _dismiss_popup(page)



        result = await page.evaluate("""

            () => {

                const body = document.body.innerText || '';

                const dateRe = /\\d{4}[-.]\\d{2}[-.]\\d{2}/g;



                let summary = '';

                const STOP_RE = /심사진행|입법진행|의안번호|제안자|\\n{3,}/;



                for (const keyword of ['제안이유', '주요내용']) {

                    const idx = body.indexOf(keyword);

                    if (idx === -1) continue;

                    let chunk = body.slice(idx + keyword.length).trimStart().slice(0, 600);

                    const stop = chunk.search(STOP_RE);

                    if (stop > 30) chunk = chunk.slice(0, stop);

                    chunk = chunk.replace(/\\s+/g, ' ').trim();

                    if (chunk.length > summary.length) summary = chunk;

                    if (summary.length >= 80) break;

                }

                summary = summary.slice(0, 200);



                let latestDate = '';

                const progIdx = body.search(/심사진행|입법진행|심사경과/);

                const searchArea = progIdx > -1

                    ? body.slice(progIdx, progIdx + 1500)

                    : body;

                const dates = searchArea.match(dateRe) || [];

                const today = new Date().toISOString().slice(0, 10);

                for (const d of dates) {

                    const norm = d.replace(/\\./g, '-');

                    if (norm <= today && norm > latestDate) latestDate = norm;

                }



                let notice = '';

                const candidates = document.querySelectorAll(

                    'span, em, strong, b, .badge, [class*="notice"], [class*="label"], [class*="state"]'

                );

                for (const el of candidates) {

                    const txt = (el.innerText || '').trim();

                    if (txt.includes('입법예고') && /\\d{4}.*~.*\\d{4}/.test(txt) && txt.length < 120) {

                        notice = txt;

                        break;

                    }

                }



                if (!notice) {

                    const m = body.match(/입법예고[중]?\\s*\\(\\s*\\d{4}[-.\\d]+\\s*~\\s*\\d{4}[-.\\d]+\\s*\\)/);

                    if (m) notice = m[0];

                }



                // 소관위원회 추출 — th/td 인접 셀 방식 우선
                let committee = '';
                const labelCells = document.querySelectorAll('th, td, dt, li');
                for (const cell of labelCells) {
                    const txt = (cell.innerText || '').trim();
                    if (txt === '소관위원회' || txt === '소관위') {
                        // 같은 tr 안의 다음 td
                        let next = cell.nextElementSibling;
                        while (next && next.tagName !== 'TD') next = next.nextElementSibling;
                        if (next) { committee = (next.innerText || '').trim().slice(0, 50); break; }
                        // 부모 tr의 마지막 td
                        const tr = cell.closest('tr');
                        if (tr) {
                            const tds = tr.querySelectorAll('td');
                            if (tds.length) { committee = (tds[tds.length - 1].innerText || '').trim().slice(0, 50); break; }
                        }
                    }
                }
                // 폴백1: tr 전체 텍스트에서 인접 단어 추출
                if (!committee) {
                    for (const tr of document.querySelectorAll('table tr')) {
                        const txt = (tr.innerText || '').replace(/\\s+/g, ' ').trim();
                        if (txt.includes('소관위원회')) {
                            const m = txt.match(/소관위원회\\s*([가-힣\\s]+위원회)/);
                            if (m) { committee = m[1].trim().slice(0, 50); break; }
                        }
                    }
                }
                // 폴백2: body 전체 텍스트 (줄바꿈 포함)
                if (!committee) {
                    const m = body.match(/소관위원회[\\s\\S]{0,10}?([가-힣]+위원회)/);
                    if (m) committee = m[1].trim();
                }

                return { summary, status_date: latestDate, notice, committee };

            }

        """)



        summary     = re.sub(r'\s+', ' ', result.get('summary', '')).strip()

        status_date = result.get('status_date', '').strip()

        notice      = re.sub(r'\s+', ' ', result.get('notice', '')).strip()

        committee   = result.get('committee', '').strip()

        return summary, status_date, notice, committee



    except Exception as e:

        print(f"  상세 페이지 오류({url[-30:]}): {e}")

        return "", "", ""





async def _get_total_count(page) -> int:

    el = await page.query_selector("#pager_count_text")

    if el:

        val = await el.get_attribute("value")

        try:

            return int(val)

        except (TypeError, ValueError):

            pass

    return 0





async def _parse_rows(page) -> list[dict]:

    rows = await page.query_selector_all("table.fix_tb tbody tr")

    results = []

    for row in rows:

        cols = await row.query_selector_all("td")

        if len(cols) < 4:

            continue

        try:

            bill_no = (await cols[0].inner_text()).strip()



            link_el = await cols[1].query_selector("a")

            if link_el:

                raw_title = (await link_el.get_attribute("title")) or ""

                bill_name = _clean_bill_name(raw_title)

                bill_id = (await link_el.get_attribute("data-bill-id")) or ""

                url = f"{DETAIL_BASE}?billId={bill_id}" if bill_id else ""

            else:

                bill_name = (await cols[1].inner_text()).strip()

                url = ""



            proposer = (await cols[2].inner_text()).strip()

            proposed_date = (await cols[3].inner_text()).strip()

            vote_date = (await cols[4].inner_text()).strip() if len(cols) > 4 else ""

            status = (await cols[7].inner_text()).strip() if len(cols) > 7 else ""



            if not _is_active(status):

                continue



            results.append({

                "bill_no": bill_no,

                "bill_name": bill_name,

                "proposer": proposer,

                "proposed_date": proposed_date,

                "vote_date": vote_date,

                "status": status,

                "url": url,

            })

        except Exception as e:

            print(f"  행 파싱 오류: {e}")

    return results





async def _wait_overlay(page, timeout: int = 10000):

    try:

        await page.wait_for_selector(

            "#glb-loading", state="hidden", timeout=timeout

        )

    except Exception:

        await asyncio.sleep(1)


async def _dismiss_popup(page):
    """popup-zone__wrap 등 팝업/오버레이를 JS로 강제 제거."""
    try:
        await page.evaluate("""
            () => {
                document.querySelectorAll(
                    '.popup-zone__wrap, .popup-zone, [class*="popup"][class*="wrap"], ' +
                    '[class*="modal-wrap"], [class*="layer-wrap"]'
                ).forEach(el => el.remove());
            }
        """)
    except Exception:
        pass


async def _safe_click(page, element):
    """팝업 제거 → 일반 클릭 → 실패 시 JS 클릭."""
    await _dismiss_popup(page)
    try:
        await element.click(timeout=10000)
    except Exception:
        try:
            await element.click(force=True, timeout=10000)
        except Exception:
            await page.evaluate("el => el.click()", element)





async def search_bills(page, keyword: str) -> list[dict]:

    print(f"\n[검색] 키워드: '{keyword}'")



    try:

        await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)

        await page.wait_for_load_state("networkidle", timeout=10000)

    except PlaywrightTimeout:

        print("  메인 페이지 타임아웃, 재시도...")

        await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60000)

        await asyncio.sleep(2)

    await _dismiss_popup(page)

    search_input = await page.query_selector("#first__input_keyword_pc")

    if not search_input:

        print("  검색 입력란 없음, 대체 셀렉터 시도...")

        search_input = await page.query_selector("input[placeholder*='의안명']")

    if not search_input:

        print("  검색 입력란을 찾을 수 없습니다.")

        return []



    await _safe_click(page, search_input)

    await search_input.fill(keyword)

    await search_input.press("Enter")



    try:

        await page.wait_for_load_state("networkidle", timeout=20000)

    except PlaywrightTimeout:

        await asyncio.sleep(3)

    await _dismiss_popup(page)



    total = await _get_total_count(page)

    if total == 0:

        print("  결과 없음")

        return []



    print(f"  총 {total}건 발견")



    all_results = []



    page_results = await _parse_rows(page)

    for r in page_results:

        r["keyword"] = keyword

    all_results.extend(page_results)

    print(f"  페이지 1: {len(page_results)}건")



    page_num = 2

    while True:

        next_link = await page.query_selector(

            f"a.number.page-number[onclick='fnSearch({page_num})']"

        )

        if not next_link:

            next_group = await page.query_selector("a.btn-next-group, a[title*='다음 페이지']")

            if not next_group:

                break

            await _wait_overlay(page)

            await _safe_click(page, next_group)

            try:

                await page.wait_for_load_state("networkidle", timeout=15000)

            except PlaywrightTimeout:

                await asyncio.sleep(2)

            await _wait_overlay(page)

            await _dismiss_popup(page)

            next_link = await page.query_selector(

                f"a.number.page-number[onclick='fnSearch({page_num})']"

            )

            if not next_link:

                break



        await _wait_overlay(page)

        await _safe_click(page, next_link)

        try:

            await page.wait_for_load_state("networkidle", timeout=15000)

        except PlaywrightTimeout:

            await asyncio.sleep(2)

        await _wait_overlay(page)

        await _dismiss_popup(page)



        page_results = await _parse_rows(page)

        if not page_results:

            break

        for r in page_results:

            r["keyword"] = keyword

        all_results.extend(page_results)

        print(f"  페이지 {page_num}: {len(page_results)}건")



        if len(all_results) >= total:

            break



        page_num += 1

        await asyncio.sleep(0.3)



    print(f"  수집 완료: {len(all_results)}건")

    return all_results





def save_results(all_results: list[dict], timestamp: str):

    """결과를 JSON과 CSV로 저장한다."""

    json_path = f"assembly_results_{timestamp}.json"

    csv_path = f"assembly_results_{timestamp}.csv"



    with open(json_path, "w", encoding="utf-8") as f:

        json.dump(all_results, f, ensure_ascii=False, indent=2)

    print(f"\n[저장] JSON: {json_path}")



    if all_results:

        fieldnames = [

            "priority_mark", "is_key_member", "keyword", "bill_no", "bill_name", "proposer",

            "proposed_date", "vote_date", "status", "status_changed", "status_change_label",

            "prev_status", "status_changed_date",

            "legislative_notice", "summary", "url",

        ]

        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:

            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")

            writer.writeheader()

            writer.writerows(all_results)

        print(f"[저장] CSV: {csv_path}")



    return json_path, csv_path





async def main():

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    all_results = []



    history = _load_history()

    print(f"[이력 DB] 기존 이력 {len(history)}건 로드")



    async with async_playwright() as p:

        browser = await p.chromium.launch(

            headless=False,

        )



        # ── 1. 키워드별 병렬 검색 ─────────────────────────────────────────

        sem_search = asyncio.Semaphore(SEARCH_CONCURRENCY)



        async def _search_one(keyword: str) -> list[dict]:

            async with sem_search:

                ctx = await browser.new_context(

                    viewport={"width": 1280, "height": 900},

                    locale="ko-KR",

                )

                pg = await ctx.new_page()

                try:

                    return await search_bills(pg, keyword)

                except Exception as e:

                    print(f"  '{keyword}' 검색 오류: {e}")

                    return []

                finally:

                    await ctx.close()



        print(f"\n[병렬 검색] {len(KEYWORDS)}개 키워드 동시 검색 (최대 {SEARCH_CONCURRENCY}개)")

        results_per_kw = await asyncio.gather(*[_search_one(kw) for kw in KEYWORDS])

        for res in results_per_kw:

            all_results.extend(res)



        # ── 중복 제거 (bill_no 기준) ──────────────────────────────────

        seen_no: set[str] = set()

        unique: list[dict] = []

        for r in all_results:

            key = r.get("bill_no") or r.get("bill_name", "")

            if key and key not in seen_no:

                seen_no.add(key)

                unique.append(r)

        all_results = unique



        # ── 2. 목록 페이지 1차 필터: 날짜 조건 통과한 것만 상세 방문 ──────

        to_detail = [r for r in all_results if _is_recent(r)]

        # history_db에 입법예고 기간 중으로 기록된 의안 → 목록에 없어도 재방문
        to_detail_keys = {r.get("bill_no") or r.get("url", "") for r in to_detail}
        for hist_no, hist_row in history.items():
            if not _is_notice_active(str(hist_row.get("legislative_notice") or "")):
                continue
            if hist_no in to_detail_keys:
                continue
            url = str(hist_row.get("url") or "")
            if not url:
                continue
            # to_detail에 복원 (목록 페이지 기본값으로)
            to_detail.append({
                "bill_no":       hist_no,
                "bill_name":     str(hist_row.get("bill_name") or ""),
                "proposer":      str(hist_row.get("proposer") or ""),
                "proposed_date": str(hist_row.get("proposed_date") or ""),
                "vote_date":     str(hist_row.get("vote_date") or ""),
                "status":        str(hist_row.get("status") or ""),
                "keyword":       str(hist_row.get("keyword") or ""),
                "url":           url,
                "legislative_notice": str(hist_row.get("legislative_notice") or ""),
                "_from_history": True,
            })
            to_detail_keys.add(hist_no)

        skipped = len(all_results) - len(to_detail)

        print(f"\n[1차 필터] {len(all_results)}건 → {len(to_detail)}건 상세 조회 "

              f"({skipped}건 발의·의결일 {RECENT_DAYS}일 이내 아님, 상세 방문 생략)")



        # ── 3. 상세 페이지 병렬 수집 ────────────────────────────────────

        print(f"[상세 페이지 수집] {len(to_detail)}건 병렬 조회 중 (최대 {DETAIL_CONCURRENCY}개)...")

        detail_ctx = await browser.new_context(

            viewport={"width": 1280, "height": 900},

            locale="ko-KR",

        )

        sem_detail = asyncio.Semaphore(DETAIL_CONCURRENCY)



        async def _fetch_detail_one(bill: dict, idx: int) -> None:

            url = bill.get("url", "")

            if not url:

                bill["_raw_summary"] = ""

                bill["status_changed_date"] = ""

                bill["legislative_notice"] = ""

                return

            async with sem_detail:

                pg = await detail_ctx.new_page()

                try:

                    print(f"  {idx}/{len(to_detail)} {bill.get('bill_name','')[:40]}")

                    raw, status_date, notice, committee = await _fetch_bill_detail(pg, url)

                finally:

                    await pg.close()

            bill["_raw_summary"] = raw

            bill["status_changed_date"] = status_date

            bill["legislative_notice"] = notice

            bill["committee"] = committee

            if notice:

                print(f"    → 입법예고 감지: {notice}")



        await asyncio.gather(*[

            _fetch_detail_one(b, i) for i, b in enumerate(to_detail, 1)

        ])

        await detail_ctx.close()



        # 상세 미방문 의안 필드 초기화

        detail_nos = {r.get("bill_no") for r in to_detail}

        for bill in all_results:

            if bill.get("bill_no") not in detail_nos:

                bill.setdefault("_raw_summary", "")

                bill.setdefault("status_changed_date", "")

                bill.setdefault("legislative_notice", "")



        await browser.close()



    # ── 4. Claude 요약 (순차 처리 — subprocess 블로킹) ───────────────────

    print(f"\n[Claude 요약] {len(to_detail)}건 요약 생성 중...")

    for bill in to_detail:

        raw = bill.pop("_raw_summary", "")

        if raw:

            print(f"  → {bill.get('bill_name','')[:40]} 요약 중...")

            bill["summary"] = _summarize_with_claude(raw, bill.get("bill_name", ""))

        else:

            bill["summary"] = ""



    for bill in all_results:

        bill.pop("_raw_summary", None)

        bill.setdefault("summary", "")



    # ── 상태변경 레이블 적용 (전체) ──────────────────────────────────────

    for bill in all_results:

        _apply_status_change_label(bill, history)



    # ── history_db.xlsx: 필터 전 전체 수집 의안 누적 저장 ────────────────

    _save_history(all_results, history)



    # ── 최종 필터: 발의일 7일 이내 OR 심사변경일 7일 이내 OR 입법예고 중 ──

    before = len(all_results)

    all_results = [

        r for r in all_results

        if _is_recent(r)

        or _is_recent_status(r.get("status_changed_date", ""))

        or _is_notice_active(r.get("legislative_notice", ""))

    ]

    print(f"\n[최종 필터] {before}건 → {len(all_results)}건 "

          f"(발의일·심사변경 7일 이내, 또는 입법예고 기간 중)")



    # ── 추가 필터: 키워드별 콘텐츠 조건 적용 ────────────────────────────────

    before_narrow = len(all_results)

    all_results = [r for r in all_results if _passes_content_filter(r)]

    if before_narrow != len(all_results):

        print(f"[필터] 콘텐츠 필터: {before_narrow}건 → {len(all_results)}건")



    # ── 우선순위 의원 ★ 표시 추가 ────────────────────────────────────

    for bill in all_results:

        if _is_priority(bill):

            bill["is_key_member"] = True

            bill["priority_mark"] = "★"

            if not bill["bill_name"].startswith("★"):

                bill["bill_name"] = "★ " + bill["bill_name"]

        else:

            bill["is_key_member"] = False

            bill["priority_mark"] = ""



    # ── 정렬: ① 우선순위 의원 ② 상태변경 의안 ③ 최신 발의일 순 ──────

    all_results.sort(

        key=lambda r: (

            0 if r.get("priority_mark") == "★" else 1,

            0 if r.get("status_change_label") else 1,

            "9999-99-99" if not r.get("proposed_date") else r["proposed_date"],

        ),

        reverse=True,

    )

    # 1·2순위(우선의원, 상태변경)는 오름차순(0이 앞), 날짜는 내림차순 → 두 단계 정렬

    all_results.sort(

        key=lambda r: (

            0 if r.get("priority_mark") == "★" else 1,

            0 if r.get("status_change_label") else 1,

        )

    )



    print("\n" + "=" * 60)

    print("검색 결과 요약")

    print("=" * 60)

    for keyword in KEYWORDS:

        count = sum(1 for r in all_results if r.get("keyword") == keyword)

        if count:

            print(f"  {keyword}: {count}건")

    priority_count = sum(1 for r in all_results if r.get("is_key_member"))

    changed_count  = sum(1 for r in all_results if r.get("status_changed"))

    print(f"  전체: {len(all_results)}건  (주요의원 ★ {priority_count}건, 의결결과변동 {changed_count}건)")



    # 의결결과 변동 의안 목록 출력

    if changed_count:

        print("\n[의결결과 변동 의안]")

        for r in all_results:

            if r.get("status_changed"):

                mark = "★ " if r.get("is_key_member") else ""

                name = r.get("bill_name", "")[:45]

                label = r.get("status_change_label", "")

                print(f"  {mark}{name}  {label}")



    # 0건이어도 반드시 저장

    save_results(all_results, timestamp)

    if not all_results:

        print("\n수집된 결과가 없습니다.")



    return all_results





if __name__ == "__main__":

    asyncio.run(main())